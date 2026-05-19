"""生成测试用 Excel fixture。"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent


def make_en_report():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Header style
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    center = Alignment(horizontal="center")

    headers = ["Product", "Category", "Revenue", "Growth", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Data rows
    rows = [
        ("Laptop Pro",     "Electronics", 128000, 0.23, "Strong demand in Q4"),
        ("Office Chair",   "Furniture",    45000, 0.08, "Supply chain delays"),
        ("Wireless Mouse", "Electronics",  32000, 0.15, "Best seller this month"),
        ("Standing Desk",  "Furniture",    67000, 0.31, "New product launch"),
        ("Monitor 4K",     "Electronics",  89000, 0.19, "Premium segment growth"),
    ]
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    # Summary row
    ws.cell(row=8, column=1, value="Total").font = Font(bold=True)
    ws.cell(row=8, column=3, value=361000).font = Font(bold=True)
    ws.cell(row=8, column=5, value="Quarterly summary").font = Font(bold=True)

    # Second sheet
    ws2 = wb.create_sheet("Team Notes")
    ws2["A1"] = "Meeting Notes"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A2"] = "Action Items"
    ws2["A3"] = "Review pricing strategy for electronics"
    ws2["A4"] = "Follow up with furniture suppliers"
    ws2["A5"] = "Prepare Q1 forecast report"

    # Column widths
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 20

    wb.save(str(OUT / "en_report.xlsx"))
    print("saved en_report.xlsx")


if __name__ == "__main__":
    make_en_report()
