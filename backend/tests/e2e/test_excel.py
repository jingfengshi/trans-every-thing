"""E2E: Excel 翻译测试"""
import pytest
from openpyxl import load_workbook
from pathlib import Path
from .conftest import FIXTURES_DIR, submit_and_wait


def test_excel_pipeline(app_client):
    client, output_dir = app_client
    result = submit_and_wait(client, FIXTURES_DIR / "en_report.xlsx", engine="claude", target_lang="zh")

    status = client.get(f"/api/status/{result['task_id']}").json()
    assert status["status"] == "done", f"Task failed: {status}"


def test_excel_output_file_exists(app_client):
    client, output_dir = app_client
    result = submit_and_wait(client, FIXTURES_DIR / "en_report.xlsx", target_lang="zh")
    task_id = result["task_id"]

    output_path = output_dir / f"{task_id}_translated.xlsx"
    assert output_path.exists(), "Output .xlsx not created"


def test_excel_output_is_valid_workbook(app_client):
    client, output_dir = app_client
    result = submit_and_wait(client, FIXTURES_DIR / "en_report.xlsx", target_lang="zh")
    output_path = output_dir / f"{result['task_id']}_translated.xlsx"

    wb = load_workbook(str(output_path))
    assert len(wb.sheetnames) == 2  # Sales Report + Team Notes
    wb.close()


def test_excel_text_translated(app_client):
    """输出文本包含翻译标记 [ZH]"""
    client, output_dir = app_client
    result = submit_and_wait(client, FIXTURES_DIR / "en_report.xlsx", target_lang="zh")
    output_path = output_dir / f"{result['task_id']}_translated.xlsx"

    wb = load_workbook(str(output_path))
    all_text = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    all_text.append(cell.value)
    wb.close()

    assert any("[ZH]" in t for t in all_text), \
        f"No translated text found. Sample: {all_text[:5]}"


def test_excel_numeric_cells_unchanged(app_client):
    """数字单元格不应被翻译"""
    client, output_dir = app_client
    result = submit_and_wait(client, FIXTURES_DIR / "en_report.xlsx", target_lang="zh")
    output_path = output_dir / f"{result['task_id']}_translated.xlsx"

    wb = load_workbook(str(output_path))
    ws = wb["Sales Report"]
    # revenue column (col 3) should still be numeric
    revenue_vals = [ws.cell(row=r, column=3).value for r in range(2, 7)]
    wb.close()

    assert all(isinstance(v, (int, float)) for v in revenue_vals if v is not None), \
        f"Numeric cells were modified: {revenue_vals}"


def test_excel_sheet_names_preserved(app_client):
    """Sheet 名称保留"""
    client, output_dir = app_client
    result = submit_and_wait(client, FIXTURES_DIR / "en_report.xlsx", target_lang="zh")
    output_path = output_dir / f"{result['task_id']}_translated.xlsx"

    orig = load_workbook(str(FIXTURES_DIR / "en_report.xlsx"))
    out  = load_workbook(str(output_path))
    assert out.sheetnames == orig.sheetnames
    orig.close()
    out.close()


def test_excel_download_endpoint(app_client):
    client, output_dir = app_client
    result = submit_and_wait(client, FIXTURES_DIR / "en_report.xlsx", target_lang="zh")
    task_id = result["task_id"]

    resp = client.get(f"/api/download/{task_id}")
    assert resp.status_code == 200
    assert "spreadsheet" in resp.headers["content-type"]
