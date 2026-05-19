import copy
import shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from .base import BaseProcessor, TextBlock


class ExcelProcessor(BaseProcessor):
    """
    Excel processor using openpyxl.
    TextBlock metadata keys:
      sheet_name: str
      row: int (1-based)
      col: int (1-based)
    x/y/width/height filled with 0 (unused for Excel).
    """

    def extract_blocks(self, file_path: str) -> list[TextBlock]:
        wb = load_workbook(file_path, data_only=True)
        blocks = []
        page = 0  # use page index as sheet index

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    text = str(cell.value).strip()
                    if not text:
                        continue
                    # skip purely numeric cells — no translation needed
                    try:
                        float(text.replace(",", "").replace("%", ""))
                        continue
                    except ValueError:
                        pass

                    blocks.append(TextBlock(
                        text=text,
                        x=0, y=0, width=0, height=0,
                        font_size=cell.font.size or 11,
                        font_name=cell.font.name or "Calibri",
                        page=page,
                        metadata={
                            "sheet_name": sheet_name,
                            "row": cell.row,
                            "col": cell.column,
                        }
                    ))
            page += 1

        wb.close()
        return blocks

    def rebuild(self, original_path: str, blocks: list[TextBlock], output_path: str) -> None:
        # copy original to preserve all styles, formulas, images
        shutil.copy2(original_path, output_path)
        wb = load_workbook(output_path)

        # index blocks by (sheet_name, row, col)
        lookup: dict[tuple, str] = {}
        for block in blocks:
            key = (block.metadata["sheet_name"], block.metadata["row"], block.metadata["col"])
            lookup[key] = block.text

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    key = (sheet_name, cell.row, cell.column)
                    if key in lookup:
                        cell.value = lookup[key]

        wb.save(output_path)
        wb.close()
